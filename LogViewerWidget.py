# -*- coding: utf-8 -*-
"""
日志查看器Widget
用于查看、搜索和打开日志文件
"""

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QLineEdit, QPushButton, QListWidget, QTextEdit,
                             QFileDialog, QMessageBox, QGroupBox, QDateEdit,
                             QTableWidget, QTableWidgetItem, QHeaderView,
                             QDialog, QRadioButton, QButtonGroup)
from PyQt5.QtCore import Qt, QDate, QDateTime
from PyQt5.QtGui import QFont
import os
import glob
import traceback
from datetime import datetime
from openpyxl import load_workbook

from src.support.log_utils import parse_log_excel, merge_log_infos, write_log_to_excel
from src.support.support_funs import sanitize_filename_part


class BatchSummaryDialog(QDialog):
    """生成批次总表对话框：输入 lot_id，选择工位"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("生成批次总表")
        self.lot_id = ""
        self.station_filter = "all"
        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Lot ID:"))
        self.lot_id_input = QLineEdit()
        self.lot_id_input.setPlaceholderText("输入 lot_id...")
        layout.addWidget(self.lot_id_input)

        layout.addWidget(QLabel("工位:"))
        btn_group = QButtonGroup(self)
        self.radio_dry = QRadioButton("仅 dry")
        self.radio_transfer = QRadioButton("仅 transfer")
        self.radio_all = QRadioButton("全部 (dry + transfer)")
        self.radio_all.setChecked(True)
        btn_group.addButton(self.radio_dry)
        btn_group.addButton(self.radio_transfer)
        btn_group.addButton(self.radio_all)
        layout.addWidget(self.radio_dry)
        layout.addWidget(self.radio_transfer)
        layout.addWidget(self.radio_all)

        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("确定")
        cancel_btn = QPushButton("取消")
        ok_btn.clicked.connect(self._on_ok)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

    def _on_ok(self):
        raw = self.lot_id_input.text().strip()
        if not raw:
            QMessageBox.warning(self, "校验", "请输入 lot_id")
            return
        self.lot_id = raw
        if self.radio_dry.isChecked():
            self.station_filter = "dry"
        elif self.radio_transfer.isChecked():
            self.station_filter = "transfer"
        else:
            self.station_filter = "all"
        self.accept()


class LogViewerWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.log_dir = os.path.abspath("Log")
        
        self.init_ui()
        self._update_folder_display()
        self.refresh_log_list()
    
    def init_ui(self):
        """初始化UI"""
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # 搜索区域
        search_group = QGroupBox("搜索")
        search_layout = QHBoxLayout()
        
        # 文件名搜索
        self.search_label = QLabel("文件名搜索:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入文件名关键字...")
        self.search_input.textChanged.connect(self.filter_logs)
        
        # 日期搜索
        self.date_label = QLabel("日期搜索:")
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setDate(QDate.currentDate())
        self.date_input.setDisplayFormat("yyyy-MM-dd")
        self.date_input.dateChanged.connect(self.filter_logs)
        
        # 清除搜索按钮
        self.clear_search_btn = QPushButton("清除搜索")
        self.clear_search_btn.clicked.connect(self.clear_search)
        
        # 刷新按钮
        self.refresh_btn = QPushButton("刷新列表")
        self.refresh_btn.clicked.connect(self.refresh_log_list)

        # 生成批次总表按钮
        self.batch_summary_btn = QPushButton("生成批次总表")
        self.batch_summary_btn.clicked.connect(self._on_batch_summary)
        
        search_layout.addWidget(self.search_label)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.date_label)
        search_layout.addWidget(self.date_input)
        search_layout.addWidget(self.clear_search_btn)
        search_layout.addWidget(self.refresh_btn)
        search_layout.addWidget(self.batch_summary_btn)
        search_layout.addStretch()
        
        search_group.setLayout(search_layout)
        main_layout.addWidget(search_group)
        
        # 文件列表和内容查看区域
        content_layout = QHBoxLayout()
        
        # 左侧：文件列表
        list_group = QGroupBox("日志文件列表")
        list_layout = QVBoxLayout()

        # 文件夹选择
        folder_layout = QHBoxLayout()
        folder_layout.addWidget(QLabel("当前目录:"))
        self.folder_path_edit = QLineEdit()
        self.folder_path_edit.setReadOnly(True)
        self.folder_path_edit.setPlaceholderText("选择日志目录...")
        folder_layout.addWidget(self.folder_path_edit, 1)
        self.select_folder_btn = QPushButton("选择文件夹")
        self.select_folder_btn.clicked.connect(self._on_select_folder)
        folder_layout.addWidget(self.select_folder_btn)
        list_layout.addLayout(folder_layout)
        
        self.log_list = QListWidget()
        self.log_list.setMinimumWidth(400)
        self.log_list.itemDoubleClicked.connect(self.open_log_file)
        self.log_list.itemSelectionChanged.connect(self.on_selection_changed)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        
        # 打开按钮
        self.open_btn = QPushButton("打开选中文件")
        self.open_btn.clicked.connect(self.open_selected_log)
        
        # 删除按钮
        self.delete_btn = QPushButton("删除选中文件")
        self.delete_btn.clicked.connect(self.delete_selected_log)
        self.delete_btn.setStyleSheet("QPushButton { background-color: #f44336; color: white; }")
        
        button_layout.addWidget(self.open_btn)
        button_layout.addWidget(self.delete_btn)
        
        list_layout.addWidget(self.log_list)
        list_layout.addLayout(button_layout)
        list_group.setLayout(list_layout)
        
        # 右侧：日志内容显示
        content_group = QGroupBox("日志内容")
        content_view_layout = QVBoxLayout()
        
        # 使用QTableWidget显示Excel内容，支持单元格居中
        self.log_table = QTableWidget()
        self.log_table.setEditTriggers(QTableWidget.NoEditTriggers)  # 只读
        self.log_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.log_table.setAlternatingRowColors(True)  # 交替行颜色
        self.log_table.horizontalHeader().setStretchLastSection(True)
        self.log_table.verticalHeader().setVisible(False)
        
        content_view_layout.addWidget(self.log_table)
        content_group.setLayout(content_view_layout)
        
        content_layout.addWidget(list_group, 1)
        content_layout.addWidget(content_group, 2)
        
        main_layout.addLayout(content_layout)
    
    def _update_folder_display(self):
        """更新文件夹路径显示"""
        self.folder_path_edit.setText(self.log_dir)

    def _on_select_folder(self):
        """选择日志文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择日志目录", self.log_dir)
        if not folder:
            return
        self.log_dir = os.path.abspath(folder)
        self._update_folder_display()
        self.log_table.setRowCount(0)
        self.log_table.setColumnCount(0)
        self.refresh_log_list()

    def refresh_log_list(self):
        """刷新日志文件列表"""
        self.log_list.clear()
        
        if not os.path.exists(self.log_dir):
            return
        
        # 获取所有.xlsx文件
        log_files = glob.glob(os.path.join(self.log_dir, "*.xlsx"))
        log_files.sort(reverse=True)  # 按时间倒序排列
        
        for log_file in log_files:
            filename = os.path.basename(log_file)
            self.log_list.addItem(filename)
        
        if self.log_list.count() > 0:
            self.log_list.setCurrentRow(0)
    
    def filter_logs(self):
        """根据搜索条件过滤日志列表"""
        search_text = self.search_input.text().lower()
        search_date = self.date_input.date().toString("yyyyMMdd")
        
        for i in range(self.log_list.count()):
            item = self.log_list.item(i)
            filename = item.text().lower()
            
            # 检查文件名是否包含搜索文本
            text_match = search_text == "" or search_text in filename
            
            # 检查日期是否匹配
            date_match = search_date in filename or search_text != ""
            
            item.setHidden(not (text_match and date_match))
    
    def clear_search(self):
        """清除搜索条件"""
        self.search_input.clear()
        self.date_input.setDate(QDate.currentDate())
        self.filter_logs()
    
    def on_selection_changed(self):
        """选中项改变时的处理"""
        current_item = self.log_list.currentItem()
        if current_item:
            self.open_btn.setEnabled(True)
            self.delete_btn.setEnabled(True)
        else:
            self.open_btn.setEnabled(False)
            self.delete_btn.setEnabled(False)
    
    def open_selected_log(self):
        """打开选中的日志文件"""
        current_item = self.log_list.currentItem()
        if current_item:
            self.open_log_file(current_item)
    
    def delete_selected_log(self):
        """删除选中的日志文件"""
        current_item = self.log_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "警告", "请先选择一个日志文件")
            return
        
        filename = current_item.text()
        filepath = os.path.join(self.log_dir, filename)
        
        # 确认删除
        reply = QMessageBox.question(
            self, 
            "确认删除", 
            f"确定要删除日志文件吗？\n\n文件名: {filename}\n\n此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    QMessageBox.information(self, "成功", f"已成功删除文件: {filename}")
                    # 清空表格显示
                    self.log_table.setRowCount(0)
                    self.log_table.setColumnCount(0)
                    # 刷新列表
                    self.refresh_log_list()
                else:
                    QMessageBox.warning(self, "错误", f"文件不存在: {filepath}")
            except Exception as e:
                error_msg = f"删除文件失败: {str(e)}\n"
                error_msg += f"文件路径: {filepath}\n"
                error_msg += f"错误详情: {traceback.format_exc()}"
                QMessageBox.critical(self, "错误", error_msg)
    
    def open_log_file(self, item):
        """打开日志文件并显示内容"""
        if item is None:
            return
        filename = item.text()
        filepath = os.path.join(self.log_dir, filename)
        self._display_excel_content(filepath)

    def _display_excel_content(self, filepath: str):
        """将 Excel 文件内容显示到右侧表格"""
        if not os.path.exists(filepath):
            QMessageBox.warning(self, "错误", f"文件不存在: {filepath}")
            return
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            max_row = ws.max_row
            max_col = ws.max_column

            header_items = []
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            for col_idx, header_value in enumerate(first_row):
                if header_value is not None:
                    header_items.append(str(header_value))
                else:
                    header_items.append(f"列{col_idx + 1}")

            data_row_count = max_row - 1 if max_row > 1 else 0
            self.log_table.setRowCount(data_row_count)
            self.log_table.setColumnCount(max_col)
            self.log_table.setHorizontalHeaderLabels(header_items)

            if data_row_count > 0:
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
                    for col_idx, cell_value in enumerate(row, start=0):
                        if col_idx >= max_col:
                            break
                        cell_str = str(cell_value) if cell_value is not None else ""
                        tw_item = QTableWidgetItem(cell_str)
                        tw_item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                        self.log_table.setItem(row_idx, col_idx, tw_item)

            self.log_table.resizeColumnsToContents()
            for col_idx in range(max_col):
                if self.log_table.columnWidth(col_idx) < 80:
                    self.log_table.setColumnWidth(col_idx, 80)
            self.log_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)

        except Exception as e:
            error_msg = f"读取文件错误: {str(e)}\n文件路径: {filepath}\n{traceback.format_exc()}"
            QMessageBox.critical(self, "错误", error_msg)
            self.log_table.setRowCount(1)
            self.log_table.setColumnCount(1)
            self.log_table.setItem(0, 0, QTableWidgetItem(error_msg))

    def _on_batch_summary(self):
        """生成批次总表"""
        dlg = BatchSummaryDialog(self)
        if dlg.exec_() != QDialog.Accepted:
            return
        lot_id = dlg.lot_id
        station_filter = dlg.station_filter

        log_dir = self.log_dir
        if not os.path.exists(log_dir):
            QMessageBox.warning(self, "提示", "未找到匹配日志")
            return

        sanitized = sanitize_filename_part(lot_id)
        pattern = os.path.join(log_dir, f"{sanitized}_*.xlsx")
        candidates = glob.glob(pattern)

        # 按工位过滤，排除 batch_summary
        def match_station(fpath):
            name = os.path.basename(fpath)
            if "_batch_summary_" in name:
                return False
            if station_filter == "dry":
                return "_dry_" in name
            if station_filter == "transfer":
                return "_transfer_" in name
            return "_dry_" in name or "_transfer_" in name

        matched = [p for p in candidates if match_station(p)]

        # 按文件名时间戳升序（提取 _YYYYMMDD_HHMMSS 部分）
        def extract_ts(p):
            name = os.path.basename(p)
            parts = name.rsplit("_", 2)
            if len(parts) >= 3:
                try:
                    return parts[-2] + "_" + parts[-1].replace(".xlsx", "")
                except Exception:
                    pass
            return name

        matched.sort(key=extract_ts)

        if not matched:
            QMessageBox.warning(self, "提示", "未找到匹配日志")
            return

        log_info_list = []
        failed_files = []
        for fp in matched:
            info = parse_log_excel(fp)
            if info:
                log_info_list.append(info)
            else:
                failed_files.append(os.path.basename(fp))

        if not log_info_list:
            msg = "未找到可解析的匹配日志"
            if failed_files:
                msg += f"\n解析失败: {', '.join(failed_files[:5])}{'...' if len(failed_files) > 5 else ''}"
            QMessageBox.warning(self, "提示", msg)
            return

        merged = merge_log_infos(log_info_list, station_filter)

        summary_dir = os.path.join(log_dir, "Summary")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_filename = f"{sanitized}_batch_summary_{timestamp}.xlsx"
        out_path = os.path.join(summary_dir, out_filename)

        try:
            write_log_to_excel(merged, out_path)
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败: {str(e)}\n{traceback.format_exc()}")
            return

        success_msg = f"批次总表已保存: {out_filename}"
        if failed_files:
            success_msg += f"\n\n共 {len(matched)} 个文件，{len(log_info_list)} 个解析成功。"
            success_msg += f"\n解析失败: {', '.join(failed_files[:3])}{'...' if len(failed_files) > 3 else ''}"
        QMessageBox.information(self, "成功", success_msg)
        self._display_excel_content(out_path)
