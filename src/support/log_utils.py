# -*- coding: utf-8 -*-
"""
日志解析、合并与写入工具
用于批次总表整合功能
"""

from __future__ import annotations

import os
import re
import traceback
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.support.support_funs import sanitize_filename_part

# Excel 中检测类型标签到键名的反向映射
LABEL_TO_DEFECT_KEY = {
    "尺寸检测不良数": "Size",
    "锡球面积检测不良数": "Ball_Area",
    "锡球数量检测不良数": "Ball Count",
    "Mark检测不良数": "Mark",
    "划痕检测不良数": "Scratch",
    "偏移检测不良数": "Shift",
}

PRODUCT_HEADERS = ["序号", "宽度(mm)", "高度(mm)", "有Mark", "NG球数", "NG球信息", "偏移量X(mm)", "偏移量Y(mm)", "是否NG", "NG类型"]


def parse_log_excel(filepath: str) -> dict | None:
    """
    解析单条日志 Excel，返回与 get_log_info() 结构一致的 dict。
    解析失败返回 None。
    """
    if not os.path.exists(filepath):
        return None
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        max_row = ws.max_row
        if max_row < 2:
            return None

        process_info = {}
        statistics = {}
        defect_statistics = {}
        product_list = []
        product_header_row = None

        row = 1
        section = None  # "process" | "statistics" | "defect" | "product"

        while row <= max_row:
            c1 = ws.cell(row=row, column=1).value
            c2 = ws.cell(row=row, column=2).value
            c1_str = str(c1).strip() if c1 is not None else ""

            if c1_str == "检测流程信息":
                section = "process"
                row += 1
                continue
            if c1_str == "统计信息":
                section = "statistics"
                row += 1
                continue
            if c1_str == "各检测项目不良统计":
                section = "defect"
                row += 1
                continue

            # 检查是否为产品明细表头
            first_row_vals = [ws.cell(row=row, column=c).value for c in range(1, 11)]
            first_row_strs = [str(v).strip() if v is not None else "" for v in first_row_vals]
            if first_row_strs[:3] == ["序号", "宽度(mm)", "高度(mm)"] or (
                len(first_row_strs) >= 2 and first_row_strs[0] == "序号" and first_row_strs[1] == "宽度(mm)"
            ):
                product_header_row = row
                section = "product"
                row += 1
                break

            if section == "process":
                if c1_str == "开始时间":
                    process_info["start_time"] = str(c2) if c2 is not None else ""
                elif c1_str == "结束时间":
                    process_info["end_time"] = str(c2) if c2 is not None else ""
                elif c1_str == "持续时间(秒)":
                    try:
                        process_info["duration_seconds"] = float(c2)
                    except (TypeError, ValueError):
                        pass
                elif c1_str == "总产品数":
                    try:
                        process_info["total_products"] = int(c2) if c2 is not None else 0
                    except (TypeError, ValueError):
                        process_info["total_products"] = 0
                elif c1_str == "NG总数":
                    try:
                        process_info["ng_total"] = int(c2) if c2 is not None else 0
                    except (TypeError, ValueError):
                        process_info["ng_total"] = 0
                elif c1_str == "已启用检测项目":
                    val = str(c2) if c2 else ""
                    process_info["enabled_detections"] = [x.strip() for x in val.replace("、", ",").split(",") if x.strip()] if val else []

            elif section == "statistics":
                if c1_str and c1_str not in ("", "各检测项目不良统计"):
                    # 解析统计项
                    if "宽度极值" in c1_str and c2:
                        try:
                            m = re.search(r"([\d.]+)\s*\(最大:[\d.]+\s*-\s*最小:[\d.]+\)", str(c2))
                            if m:
                                statistics["width_range"] = float(m.group(1))
                        except (ValueError, AttributeError):
                            pass
                    elif "高度极值" in c1_str and c2:
                        try:
                            m = re.search(r"([\d.]+)\s*\(最大:[\d.]+\s*-\s*最小:[\d.]+\)", str(c2))
                            if m:
                                statistics["height_range"] = float(m.group(1))
                        except (ValueError, AttributeError):
                            pass
                    elif "平均尺寸" in c1_str and c2:
                        try:
                            m = re.search(r"宽度:([\d.]+),\s*高度:([\d.]+)", str(c2))
                            if m:
                                statistics["avg_width"] = float(m.group(1))
                                statistics["avg_height"] = float(m.group(2))
                        except (ValueError, AttributeError):
                            pass
                    elif "平均球半径" in c1_str and c2:
                        try:
                            statistics["avg_ball_radius"] = float(c2)
                        except (TypeError, ValueError):
                            pass
                    elif "偏移X极值" in c1_str and c2:
                        try:
                            m = re.search(r"最大:([\d.-]+),\s*最小:([\d.-]+)", str(c2))
                            if m:
                                statistics["shift_x_max"] = float(m.group(1))
                                statistics["shift_x_min"] = float(m.group(2))
                        except (ValueError, AttributeError):
                            pass
                    elif "偏移Y极值" in c1_str and c2:
                        try:
                            m = re.search(r"最大:([\d.-]+),\s*最小:([\d.-]+)", str(c2))
                            if m:
                                statistics["shift_y_max"] = float(m.group(1))
                                statistics["shift_y_min"] = float(m.group(2))
                        except (ValueError, AttributeError):
                            pass
                    elif "偏移X CPK" in c1_str and c2 is not None:
                        try:
                            statistics["shift_x_cpk"] = float(c2)
                        except (TypeError, ValueError):
                            pass
                    elif "偏移Y CPK" in c1_str and c2 is not None:
                        try:
                            statistics["shift_y_cpk"] = float(c2)
                        except (TypeError, ValueError):
                            pass

            elif section == "defect":
                key = LABEL_TO_DEFECT_KEY.get(c1_str)
                if key is not None and c2 is not None:
                    try:
                        defect_statistics[key] = int(c2)
                    except (TypeError, ValueError):
                        defect_statistics[key] = 0

            row += 1

        # 解析产品明细
        if product_header_row is not None:
            for r in range(product_header_row + 1, max_row + 1):
                idx = ws.cell(row=r, column=1).value
                w = ws.cell(row=r, column=2).value
                h = ws.cell(row=r, column=3).value
                has_mark = ws.cell(row=r, column=4).value
                ng_ball_count = ws.cell(row=r, column=5).value
                ng_ball_info = ws.cell(row=r, column=6).value
                shift_x = ws.cell(row=r, column=7).value
                shift_y = ws.cell(row=r, column=8).value
                is_ng = ws.cell(row=r, column=9).value
                ng_types = ws.cell(row=r, column=10).value

                # 跳过空行
                if idx is None and w is None and h is None:
                    continue

                product_list.append({
                    "product_index": str(idx) if idx is not None else "",
                    "width": str(w) if w is not None else "",
                    "height": str(h) if h is not None else "",
                    "has_mark": str(has_mark) if has_mark is not None else "否",
                    "ng_ball_count": str(ng_ball_count) if ng_ball_count is not None else "0",
                    "ng_ball_info": str(ng_ball_info) if ng_ball_info is not None else "",
                    "shift_x": str(shift_x) if shift_x is not None else "0.0000",
                    "shift_y": str(shift_y) if shift_y is not None else "0.0000",
                    "is_ng": str(is_ng) if is_ng is not None else "否",
                    "ng_types": str(ng_types) if ng_types is not None else "",
                })

        return {
            "lot_id": "",
            "sn_id": "",
            "process_info": process_info,
            "statistics": statistics,
            "defect_statistics": defect_statistics,
            "product_list": product_list,
        }
    except Exception as e:
        print(f"解析日志文件失败 {filepath}: {e}")
        traceback.print_exc()
        return None


def _compute_statistics_from_product_list(product_list: list) -> dict:
    """基于 product_list 重新计算 statistics。"""
    width_values = []
    height_values = []
    shift_x_list = []
    shift_y_list = []

    for p in product_list:
        w = p.get("width", "")
        h = p.get("height", "")
        if w:
            try:
                width_values.append(float(w))
            except (ValueError, TypeError):
                pass
        if h:
            try:
                height_values.append(float(h))
            except (ValueError, TypeError):
                pass
        sx = p.get("shift_x", "0")
        sy = p.get("shift_y", "0")
        try:
            shift_x_list.append(float(sx))
        except (ValueError, TypeError):
            pass
        try:
            shift_y_list.append(float(sy))
        except (ValueError, TypeError):
            pass

    stats = {}
    if width_values:
        stats["width_range"] = max(width_values) - min(width_values)
        stats["avg_width"] = sum(width_values) / len(width_values)
    if height_values:
        stats["height_range"] = max(height_values) - min(height_values)
        stats["avg_height"] = sum(height_values) / len(height_values)
    if shift_x_list:
        stats["shift_x_max"] = max(shift_x_list)
        stats["shift_x_min"] = min(shift_x_list)
    if shift_y_list:
        stats["shift_y_max"] = max(shift_y_list)
        stats["shift_y_min"] = min(shift_y_list)

    # shift_x_cpk、shift_y_cpk、avg_ball_radius 采用各日志的平均值，不在此重算
    return stats


def merge_log_infos(log_info_list: list, station_filter: str = "all") -> dict:
    """
    合并多个 log_info。station_filter 仅用于调用方过滤，此处假定已过滤。
    返回合并后的 log_info。
    """
    if not log_info_list:
        return {
            "lot_id": "",
            "sn_id": "",
            "process_info": {},
            "statistics": {},
            "defect_statistics": {},
            "product_list": [],
        }

    # process_info 合并
    total_products = 0
    ng_total = 0
    start_times = []
    end_times = []
    enabled_set = set()

    for info in log_info_list:
        pi = info.get("process_info", {})
        total_products += pi.get("total_products", 0)
        ng_total += pi.get("ng_total", 0)
        if pi.get("start_time"):
            try:
                start_times.append(datetime.strptime(pi["start_time"], "%Y-%m-%d %H:%M:%S"))
            except (ValueError, TypeError):
                pass
        if pi.get("end_time"):
            try:
                end_times.append(datetime.strptime(pi["end_time"], "%Y-%m-%d %H:%M:%S"))
            except (ValueError, TypeError):
                pass
        for d in pi.get("enabled_detections", []):
            enabled_set.add(d)

    start_time_str = ""
    end_time_str = ""
    duration_seconds = None
    if start_times:
        start_dt = min(start_times)
        start_time_str = start_dt.strftime("%Y-%m-%d %H:%M:%S")
    if end_times:
        end_dt = max(end_times)
        end_time_str = end_dt.strftime("%Y-%m-%d %H:%M:%S")
        if start_times:
            duration_seconds = (end_dt - min(start_times)).total_seconds()

    merged_process_info = {
        "start_time": start_time_str,
        "end_time": end_time_str,
        "duration_seconds": duration_seconds,
        "total_products": total_products,
        "ng_total": ng_total,
        "enabled_detections": list(enabled_set) if enabled_set else [],
    }

    # defect_statistics 合并（取并集，缺失键按 0）
    all_keys = set()
    for info in log_info_list:
        all_keys.update(info.get("defect_statistics", {}).keys())
    defect_keys = ["Size", "Ball_Area", "Ball Count", "Mark", "Scratch", "Shift"]
    all_keys.update(defect_keys)

    merged_defect = {}
    for k in all_keys:
        merged_defect[k] = sum(info.get("defect_statistics", {}).get(k, 0) for info in log_info_list)

    # product_list 按顺序合并（调用方已按时间戳排序）
    merged_products = []
    for info in log_info_list:
        merged_products.extend(info.get("product_list", []))

    # statistics 基于合并后的 product_list 重算
    merged_statistics = _compute_statistics_from_product_list(merged_products)

    # shift_x_cpk、shift_y_cpk、avg_ball_radius 采用各日志的平均值
    for key in ("shift_x_cpk", "shift_y_cpk", "avg_ball_radius"):
        vals = [x.get("statistics", {}).get(key) for x in log_info_list]
        vals = [v for v in vals if v is not None]
        if vals:
            merged_statistics[key] = sum(vals) / len(vals)

    lot_id = log_info_list[0].get("lot_id", "") or ""
    for info in log_info_list:
        lid = info.get("lot_id", "")
        if lid:
            lot_id = lid
            break

    return {
        "lot_id": lot_id,
        "sn_id": "batch",
        "process_info": merged_process_info,
        "statistics": merged_statistics,
        "defect_statistics": merged_defect,
        "product_list": merged_products,
    }


def write_log_to_excel(log_info: dict, filepath: str, detection_enable_map: dict = None):
    """
    将 log_info 写入 Excel，格式与 dry_thread.write_log_to_file 一致。
    detection_enable_map 默认全 True；若提供则使用。
    """
    if not log_info:
        raise ValueError("log_info 为空")

    detection_type_map = {
        "Size": "尺寸检测不良数",
        "Ball_Area": "锡球面积检测不良数",
        "Ball Count": "锡球数量检测不良数",
        "Mark": "Mark检测不良数",
        "Scratch": "划痕检测不良数",
        "Shift": "偏移检测不良数",
    }
    if detection_enable_map is None:
        detection_enable_map = {k: True for k in detection_type_map}

    wb = Workbook()
    ws = wb.active
    ws.title = "检测日志"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    ng_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ng_font = Font(bold=True)
    normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    row_num = 1
    process_info = log_info.get("process_info", {})
    product_list = log_info.get("product_list", [])
    statistics = log_info.get("statistics", {})

    ws.cell(row=row_num, column=1, value="检测流程信息").font = Font(bold=True, size=12)
    row_num += 1

    if process_info.get("start_time"):
        ws.cell(row=row_num, column=1, value="开始时间")
        ws.cell(row=row_num, column=2, value=process_info["start_time"])
        row_num += 1

    ws.cell(row=row_num, column=1, value="结束时间")
    ws.cell(row=row_num, column=2, value=process_info.get("end_time", ""))
    row_num += 1

    if process_info.get("duration_seconds") is not None:
        ws.cell(row=row_num, column=1, value="持续时间(秒)")
        ws.cell(row=row_num, column=2, value=f"{process_info['duration_seconds']:.2f}")
        row_num += 1

    ws.cell(row=row_num, column=1, value="总产品数")
    ws.cell(row=row_num, column=2, value=process_info.get("total_products", 0))
    row_num += 1

    ws.cell(row=row_num, column=1, value="NG总数")
    cell = ws.cell(row=row_num, column=2, value=process_info.get("ng_total", 0))
    cell.fill = ng_fill
    cell.font = ng_font
    row_num += 1

    enabled_detections = process_info.get("enabled_detections", [])
    ws.cell(row=row_num, column=1, value="已启用检测项目")
    ws.cell(row=row_num, column=2, value="、".join(enabled_detections) if enabled_detections else "无")
    row_num += 2

    ws.cell(row=row_num, column=1, value="统计信息").font = Font(bold=True, size=12)
    row_num += 1

    width_values = []
    height_values = []
    for p in product_list:
        for k, extract in [("width", width_values), ("height", height_values)]:
            v = p.get(k, "")
            if v:
                try:
                    extract.append(float(v))
                except (ValueError, TypeError):
                    pass

    stats_info = []
    if statistics.get("width_range") is not None and width_values:
        stats_info.append(("宽度极值(mm)", f"{statistics['width_range']:.4f} (最大:{max(width_values):.4f} - 最小:{min(width_values):.4f})"))
    if statistics.get("height_range") is not None and height_values:
        stats_info.append(("高度极值(mm)", f"{statistics['height_range']:.4f} (最大:{max(height_values):.4f} - 最小:{min(height_values):.4f})"))
    if statistics.get("avg_width") is not None and statistics.get("avg_height") is not None:
        stats_info.append(("平均尺寸(mm)", f"宽度:{statistics['avg_width']:.4f}, 高度:{statistics['avg_height']:.4f}"))
    if statistics.get("avg_ball_radius") is not None:
        stats_info.append(("平均球半径(mm)", f"{statistics['avg_ball_radius']:.4f}"))
    if statistics.get("shift_x_max") is not None and statistics.get("shift_x_min") is not None:
        stats_info.append(("偏移X极值(mm)", f"最大:{statistics['shift_x_max']:.4f}, 最小:{statistics['shift_x_min']:.4f}"))
    if statistics.get("shift_y_max") is not None and statistics.get("shift_y_min") is not None:
        stats_info.append(("偏移Y极值(mm)", f"最大:{statistics['shift_y_max']:.4f}, 最小:{statistics['shift_y_min']:.4f}"))
    if statistics.get("shift_x_cpk") is not None:
        stats_info.append(("偏移X CPK", f"{statistics['shift_x_cpk']:.4f}"))
    if statistics.get("shift_y_cpk") is not None:
        stats_info.append(("偏移Y CPK", f"{statistics['shift_y_cpk']:.4f}"))

    stats_info.extend([("", ""), ("各检测项目不良统计", "")])

    for label, value in stats_info:
        ws.cell(row=row_num, column=1, value=label)
        cell = ws.cell(row=row_num, column=2, value=value)
        if label == "各检测项目不良统计":
            cell.font = Font(bold=True)
        elif label in detection_type_map.values() and isinstance(value, (int, float)) and value > 0:
            cell.fill = ng_fill
            cell.font = ng_font
        row_num += 1

    defect_statistics = log_info.get("defect_statistics", {})
    for ng_type, count in defect_statistics.items():
        if detection_enable_map.get(ng_type, True):
            label = detection_type_map.get(ng_type, ng_type)
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=count)
            if count > 0:
                cell.fill = ng_fill
                cell.font = ng_font
            row_num += 1

    row_num += 1

    headers = ["序号", "宽度(mm)", "高度(mm)", "有Mark", "NG球数", "NG球信息", "偏移量X(mm)", "偏移量Y(mm)", "是否NG", "NG类型"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
    row_num += 1

    for product in product_list:
        is_ng_product = product.get("is_ng", "否") == "是"
        row_data = [
            str(product.get("product_index", "")),
            product.get("width", ""),
            product.get("height", ""),
            product.get("has_mark", "否"),
            product.get("ng_ball_count", "0"),
            product.get("ng_ball_info", ""),
            product.get("shift_x", "0.0000"),
            product.get("shift_y", "0.0000"),
            product.get("is_ng", "否"),
            product.get("ng_types", ""),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = ng_fill if is_ng_product else normal_fill
            cell.font = ng_font if is_ng_product else None
            cell.alignment = center_alignment
        row_num += 1

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(header)), max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, row_num)))
        ws.column_dimensions[col_letter].width = min(max_len + 6, 60)

    out_dir = os.path.dirname(filepath)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(filepath)
